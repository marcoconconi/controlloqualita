import csv
import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.legend import Legend
from django.core.exceptions import ObjectDoesNotExist
from controlloqualita.services.rules import calcola_punteggio_dynamic

from openpyxl import load_workbook
import zipfile
from lxml import etree
import tempfile
import os
from django.http import HttpResponse
from controlloqualita.utils import abilita_assi_chart

# --- costanti colonne (se proprio le vuoi a indice, altrimenti passa a DictReader) ---
COL_CF_RICH = 13
COL_CF_ATT  = 14
COL_SERV    = 4
COL_CLI     = 2
COL_OP      = 11
COL_DEC     = 20
COL_TEL     = 22
COL_NOTEIN  = 130
COL_INS     = 6
COL_EVA     = 7
COL_SCA     = 134
COL_CALL1   = 85
COL_LAV     = 41
COL_PEN     = 54
COL_BAN     = 70

SERVICE_TO_CATEGORY = {
    'DOSSIER RECUPERO COMPLETO PF': 'DOSSIER_BACC',
    'PERFORMANCE REPORTING PF':      'DOSSIER_BACC',
    'RAPPORTO COMPLETO PLUS':        'DOSSIER_BACC',
    'DOSSIER RECUPERO PF BASE':      'DOSSIER_PL',
    'DOSSIER RECUPERO PF':           'DOSSIER_PL',
    'DOSSIER RECUPERO UGC':          'DOSSIER_PL',
    'RINTRACCIO POSTO LAVORO':       'DOSSIER_PL',
    'RINTRACCIO EREDI':              'EREDI',
    'ACCETTAZIONE EREDI':            'EREDI',
    'RINTRACCIO INDIRIZZO':          'RINTRACCI',
    'CERTIFICAT':                    'CERTIFICATI',
}

def get_categoria_from_servizio(servizio):
    for pattern, cat in SERVICE_TO_CATEGORY.items():
        if pattern in servizio:
            return cat
    return 'N/A'

def get_categoria(servizio):
    for pat, cat in SERVICE_TO_CATEGORY.items():
        if pat in servizio:
            return cat
    return 'N/A'


def parse_aaaammgg(s):
    s = s.strip()
    if len(s) != 8: return None
    try:
        return datetime.date(int(s[:4]), int(s[4:6]), int(s[6:8]))
    except:
        return None


def parse_ddmmyyyy_hhmmss(s):
    s = s.strip()
    if not s: return None
    try:
        day, month, year = s.split()[0].split('/')
        return datetime.date(int(year), int(month), int(day))
    except:
        return None


def calcola_punteggio(SOG):
    cat = SOG['categoria']
    D = SOG['dettagli']
    # decidi quale “condition” applicare
    if cat == 'DOSSIER_BACC':
        if D['decesso_flag'] == 'SI':
            cond = 'decesso'
        elif (D['posto_lavoro'] or D['pensione']) and D['banca']:
            cond = 'lavoro_banca'
        elif (D['posto_lavoro'] or D['pensione']):
            cond = 'lavoro_no_banca'
        elif not (D['posto_lavoro'] or D['pensione'] or D['banca']) \
             and D['telefono'] and 'contattato' in D['note_interne'].lower():
            cond = 'contattato'
        else:
            cond = 'default'

    elif cat == 'DOSSIER_PL':
        # simile a DOSSIER_BACC ma salta P1
        if D['decesso_flag'] == 'SI':
            cond = 'decesso'
        elif (D['posto_lavoro'] or D['pensione']):
            cond = 'lavoro_banca'   # per PL usi sempre P2
        elif D['telefono'] and 'contattato' in D['note_interne'].lower():
            cond = 'contattato'
        else:
            cond = 'default'

    elif cat == 'EREDI':
        cond = 'eredi_chiamato' if D['erede'] else 'eredi_default'

    elif cat == 'RINTRACCI':
        cond = 'rintracci_pos' if (D['residenza_indirizzo'] or D['decesso_flag']=='SI')\
               else 'rintracci_neg'

    else:
        cond = 'default'

    # ora leggi dal DB
    from .models import ScoringRule
    try:
        rule = ScoringRule.objects.get(category=cat, condition=cond)
        return rule.score_letter, rule.score_value
    except ObjectDoesNotExist:
        # fallback se manca la riga
        return ('N2', 3)
    
def calcola_punteggio_old(detail):
    cat = detail['categoria']
    df = detail['dettagli']['decesso_flag']
    lav = detail['dettagli']['posto_lavoro']
    pen = detail['dettagli']['pensione']
    ban = detail['dettagli']['banca']
    tel = detail['dettagli']['telefono']
    note = detail['dettagli']['note_interne'].lower()
    erd = detail['dettagli']['erede']
    res = detail['dettagli']['residenza_indirizzo']

    if cat == 'DOSSIER_BACC':
        if df == 'SI': return 'P2', 10
        if (lav or pen) and ban: return 'P1', 6
        if (lav or pen): return 'P2', 10
        if not (lav or pen or ban) and tel and 'contattato' in note:
            return 'N1', 5
        return 'N2', 3

    if cat == 'DOSSIER_PL':
        if df == 'SI': return 'P2', 10
        if (lav or pen): return 'P2', 10
        if not (lav or pen) and tel and 'contattato' in note:
            return 'N1', 5
        return 'N2', 3

    if cat == 'EREDI':
        return ('P', 10) if erd else ('N', 3)

    if cat == 'RINTRACCI':
        return ('P', 10) if (res or df=='SI') else ('N', 3)

    return ('N2', 3)


def generate_quality_report(csv_path):
    """
    Legge il CSV, aggrega per (CF, cliente, servizio), calcola punteggi e KPI,
    e restituisce un BytesIO con il file Excel pronto.
    """
    # 1) lettura e raggruppamento
    soggetti = {}
    with open(csv_path, newline='', encoding='utf-8') as f:
        reader = csv.reader(f, delimiter=';', quotechar='"')
        next(reader)  # salta header
        for row in reader:
            if not row: continue
            cf = row[COL_CF_RICH].strip()
            cli = row[COL_CLI].strip()
            srv = row[COL_SERV].strip()
            key = (cf, cli, srv)
            if not key in soggetti:
                soggetti[key] = {
                    'rows': [], 'categoria': get_categoria(srv),
                    'cf': cf, 'cliente': cli, 'servizio': srv
                }
            soggetti[key]['rows'].append(row)

    # 2) costruzione dei dettagli e calcolo punteggio + raccolta KPI
    riepilogo = []
    kpi_qual = {}
    kpi_temp = {}

    for (cf, cli, srv), data in soggetti.items():
        rows = data['rows']
        cat  = data['categoria']
        # prendo la prima riga e gli eventuali dati “variabili”
        first = rows[0]
        detail = {
            'categoria': cat,
            'decesso_flag': first[COL_DEC].strip(),
            'telefono': first[COL_TEL].strip(),
            'note_interne': first[COL_NOTEIN].strip(),
            'erede': first[COL_CALL1].strip(),
            'posto_lavoro': bool(first[COL_LAV].strip()),
            'pensione': bool(first[COL_PEN].strip()),
            'banca': bool(first[COL_BAN].strip()),
            'residenza_indirizzo': first[COL_INS]  # o COL_RES
        }
        # puoi fare merge di più righe se ti serve

        # data evasion etc.
        detail['data_evasione'] = parse_aaaammgg(first[COL_EVA])
        detail['data_scadenza'] = parse_ddmmyyyy_hhmmss(first[COL_SCA])
        detail['redattore'] = first[COL_OP].upper()

        # calcolo
        let, num = calcola_punteggio(detail)

        # aggiungo al riepilogo
        riepilogo.append((first[COL_NOM], cf, detail['redattore'],
                          srv, cli, let))

        # KPI qualità
        kpi_qual.setdefault((cli, cat), []).append(num)
        # KPI tempi
        in_time = (detail['data_evasione'] <= detail['data_scadenza']) \
                  if detail['data_evasione'] and detail['data_scadenza'] else None
        kpi_temp.setdefault((cli, cat), []).append(in_time)

    # 3) creo workbook
    wb = Workbook()
    # … (identico al tuo script: foglio Riepilogo, KPI Qualità, KPI tempi, grafici) …
    # QUI build fogli esattamente come nel tuo main(), usando openpyxl.

    # 4) scrivo in BytesIO e restituisco
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out

def generate_quality_report_from_qs(qs):
    """
    Genera un XLSX con Riepilogo, KPI Qualità, KPI Tempi e KPI Grafici
    partendo da un Django QuerySet di ImportRecord.
    Ritorna un BytesIO.
    """
    # 1) Prepara i dati come liste / dizionari analoghi a quelli del tuo script.
    dati_riepilogo = []
    kpi_qualita = {}
    kpi_tempi = {}

    # Mappa punteggi testuali -> numerici
    SCORE_MAP = {'P1':6, 'P2':10, 'N1':5, 'N2':3, 'P':10, 'N':3}

    for rec in qs:
        # ricava categoria (potresti averla memorizzata in ImportRecord.raw_data o in un campo)
        servizio = rec.servizio or rec.raw_data.get('Servizio','')
        cliente  = rec.ragione_sociale_cliente or rec.raw_data.get('RagioneSocialeCliente','')
        categoria = get_categoria_from_servizio(servizio)

        # calcola il punteggio testuale e numerico
        #p_lettera, p_num = calcola_punteggio({
        p_lettera, p_num = calcola_punteggio_dynamic({
            'categoria': categoria,
            'dettagli': {
                'decesso_flag': rec.raw_data.get('DECESSO_FLAG',''),
                'posto_lavoro': bool(rec.raw_data.get('LAVORO_DATORE_RAGIONE_SOCIALE','').strip()),
                'pensione':    bool(rec.raw_data.get('PENSIONE_ENTE_RAGIONE_SOCIALE','').strip()),
                'banca':       bool(rec.raw_data.get('BANCA_RAGIONE_SOCIALE','').strip()),
                'telefono':    rec.raw_data.get('TELEFONO',''),
                'note_interne':rec.raw_data.get('NOTE_INTERNE',''),
                'erede':       rec.raw_data.get("CHIAMATO ALL'EREDITA' 1",''),
                'residenza_indirizzo': rec.raw_data.get('RESIDENZA_NEW_INDIRIZZO',''),
            }
        })

        dati_riepilogo.append([
            rec.raw_data.get('NOMINATIVO',''), 
            rec.codice_fiscale_richiesta,
            rec.operatore,
            servizio,
            cliente,
            p_lettera
        ])

        # KPI Qualità
        kpi_qualita.setdefault((cliente, categoria), []).append(p_num)

        # KPI Tempi
        di = parse_aaaammgg(rec.raw_data.get('DataInserimento','')) 
        de = parse_aaaammgg(rec.raw_data.get('DataEvasione',''))
        ds = parse_ddmmyyyy_hhmmss(rec.raw_data.get('DataScadenza',''))
        in_time = None
        if de and ds:
            in_time = de <= ds
        kpi_tempi.setdefault((cliente, categoria), []).append(in_time)

    # 2) Costruisci l’Excel
    wb = Workbook()
    # -> Riepilogo
    ws = wb.active
    ws.title = "Riepilogo"
    ws.append(["Nominativo","CF","Redattore","Servizio","Cliente","Punteggio"])
    for row in dati_riepilogo:
        ws.append(row)

    # -> KPI Qualità
    cats = ['EREDI','DOSSIER_BACC','DOSSIER_PL','RINTRACCI']
    ws_q = wb.create_sheet("KPI Qualità")
    ws_q.append(["Cliente"] + cats)
    # calcola media
    agg_q = {}
    for (cli,cat), lst in kpi_qualita.items():
        agg_q.setdefault(cli, {})[cat] = sum(lst)/len(lst)
    # ordina
    order = sorted(agg_q.items(), key=lambda x: sum(x[1].get(c,0) for c in cats))
    for cli, _ in order:
        row = [cli] + [ round(agg_q[cli].get(c,0),2) for c in cats ]
        ws_q.append(row)

    # -> KPI Tempi
    ws_t = wb.create_sheet("KPI Tempi")
    suffix = ["_numIn","_numOut","_%In","_%Out"]
    header = ["Cliente"] + [ c+s for c in cats for s in suffix ]
    ws_t.append(header)
    for cli, _ in order:
        r = [cli]
        for cat in cats:
            L = kpi_tempi.get((cli,cat),[])
            valid = [b for b in L if b is not None]
            tot = len(valid)
            in_t  = sum(1 for b in valid if b)
            out_t = sum(1 for b in valid if b is False)
            if tot:
                pi = round(in_t/tot*100,2)
                po = round(out_t/tot*100,2)
            else:
                pi = po = 0.0
            r += [in_t, out_t, pi, po]
        ws_t.append(r)
        
        
    # Grafico KPI Cliente EREDI
    
    # 1. Crea foglio dedicato per grafico eredi
    ws_eredi_filtered = wb.create_sheet("KPI EREDI GRAFICO")
    ws_eredi_filtered.sheet_state = 'hidden'
    ws_eredi_filtered.append(["Cliente", "SLA IN", "SLA OUT"])

    for row in range(2, ws_t.max_row + 1):
        cliente = ws_t.cell(row=row, column=1).value
        sla_in  = ws_t.cell(row=row, column=2).value or 0
        sla_out = ws_t.cell(row=row, column=3).value or 0
        if sla_in > 0 or sla_out > 0:
            ws_eredi_filtered.append([cliente, sla_in, sla_out])
    
    # 2. Crea grafico usando solo i dati filtrati
    chart = BarChart()
    chart.title = "KPI EREDI"
    #chart.x_axis.title = "Clienti"
    chart.y_axis.title = "Numero pratiche"
    chart.legend.position = "tr"
    #chart.width = 30   # es. 30
    chart.height = 12  # es. 15
    chart.x_axis.label_rotation = -45  # ruota di 45 gradi # lblAlgn 

    min_data_row = 2
    max_data_row = ws_eredi_filtered.max_row

    data = Reference(ws_eredi_filtered, min_col=2, max_col=3, min_row=1, max_row=max_data_row)  # Include header per i titoli
    cats = Reference(ws_eredi_filtered, min_col=1, min_row=2, max_row=max_data_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    
    '''
    chart = BarChart()
    chart.title = "KPI EREDI"
    #chart.width = 30   # es. 30
    chart.height = 12  # es. 15
    #chart.x_axis.title = "Clienti"
    #chart.x_axis.label_rotation = 45  # ruota di 45 gradi # lblAlgn 
    
    # Abilita esplicitamente gli assi
    #chart.x_axis.majorGridlines = ChartLines()
    #chart.y_axis.majorGridlines = ChartLines()
    chart.x_axis.visible = True
    chart.y_axis.visible = True

    # Supponendo che:
    #  - la colonna 1 (A) nel foglio "ws_t" contenga i nomi Cliente
    #  - la colonna 2 (B) contenga il numero di pratiche in tempo
    #  - la colonna 3 (C) contenga il numero di pratiche fuori tempo
    #  - i dati inizino dalla riga 2 (la riga 1 è intestazione)
    #  - max_data_row sia l’ultima riga con dati    
    
    min_data_row = 2
    max_data_row = ws_eredi_filtered.max_row
    #=======
    # grafico eredi
    #=======
    
    data = Reference(ws_filtered, min_col=2, max_col=3, min_row=1, max_row=max_data_row)  # Include header per i titoli
    cats = Reference(ws_filtered, min_col=1, min_row=2, max_row=max_data_row)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    # Add data (colonne B e C = SLA IN e SLA OUT)
    #chart.add_data(
    #    Reference(ws_t, min_col=2, max_col=3, min_row=1, max_row=ws_t.max_row),
    #    titles_from_data=True  # prendi titoli da prima riga
    #)
    for EREDI_row in EREDI_valid_rows:
        chart.add_data(Reference(ws_t, min_col=2, max_col=2, min_row=1, max_row=ws_t.max_row), titles_from_data=False)
        chart.add_data(Reference(ws_t, min_col=3, max_col=3, min_row=1, max_row=ws_t.max_row), titles_from_data=False)
        #chart.set_categories(Reference(ws_t, min_col=1, max_col=1, min_row=EREDI_row, max_row=EREDI_row))    
    
    chart.legend = Legend()
    chart.legend.position = "tr"  # "t" = top (sopra il grafico, vicino al titolo)
    # Altri valori validi: 'r', 'l', 'b', 't', 'tr' (destra, sinistra, basso, sopra, angolo sup. dx)
    
    # Set categories (colonna A = nomi clienti)
    chart.set_categories(
        Reference(ws_t, min_col=1, min_row=2, max_row=ws_t.max_row)
    )    
    # assegno i titoli delle serie con SeriesLabel
    chart.series[0].tx = SeriesLabel(v="SLA IN")
    chart.series[1].tx = SeriesLabel(v="SLA OUT")
    # assegno i colori
    chart.series[0].graphicalProperties.solidFill = "0000FF"  # Blu
    chart.series[1].graphicalProperties.solidFill = "FF0000"  # Rosso
    
    # Etichette dati sulle colonne (opzionale)
    #chart.dataLabels = DataLabelList()
    #chart.dataLabels.showVal = True
    
    '''
    #=======
    # Grafico KPI Cliente BACC
    #=======
    # 1. Crea foglio dedicato per grafico eredi
    ws_bacc_filtered = wb.create_sheet("KPI BACC GRAFICO")
    ws_bacc_filtered.sheet_state = 'hidden'
    ws_bacc_filtered.append(["Cliente", "SLA IN", "SLA OUT"])

    for row in range(2, ws_t.max_row + 1):
        cliente = ws_t.cell(row=row, column=1).value
        sla_in  = ws_t.cell(row=row, column=6).value or 0
        sla_out = ws_t.cell(row=row, column=7).value or 0
        if sla_in > 0 or sla_out > 0:
            ws_bacc_filtered.append([cliente, sla_in, sla_out])
    
    # 2. Crea grafico usando solo i dati filtrati
    chartBacc = BarChart()
    chartBacc.title = "KPI BACC"
    chartBacc.y_axis.title = "Numero pratiche"
    chartBacc.legend.position = "tr"
    chartBacc.height = 12  # es. 15
    chartBacc.x_axis.label_rotation = -45  # ruota di 45 gradi # lblAlgn 

    min_data_row = 2
    max_data_row = ws_bacc_filtered.max_row

    data = Reference(ws_bacc_filtered, min_col=2, max_col=3, min_row=1, max_row=max_data_row)  # Include header per i titoli
    cats = Reference(ws_bacc_filtered, min_col=1, min_row=2, max_row=max_data_row)

    chartBacc.add_data(data, titles_from_data=True)
    chartBacc.set_categories(cats)
    
    
    #=======
    # Grafico KPI Cliente DOSSIER PL
    #=======
    # 1. Crea foglio dedicato per grafico eredi
    ws_dossierpl_filtered = wb.create_sheet("KPI DOSSIER PL GRAFICO")
    ws_dossierpl_filtered.sheet_state = 'hidden'
    ws_dossierpl_filtered.append(["Cliente", "SLA IN", "SLA OUT"])

    for row in range(2, ws_t.max_row + 1):
        cliente = ws_t.cell(row=row, column=1).value
        sla_in  = ws_t.cell(row=row, column=10).value or 0
        sla_out = ws_t.cell(row=row, column=11).value or 0
        if sla_in > 0 or sla_out > 0:
            ws_dossierpl_filtered.append([cliente, sla_in, sla_out])
    
    # 2. Crea grafico usando solo i dati filtrati
    chartDossierPL = BarChart()
    chartDossierPL.title = "KPI DOSSIER PL"
    chartDossierPL.y_axis.title = "Numero pratiche"
    chartDossierPL.legend.position = "tr"
    chartDossierPL.height = 12  # es. 15
    chartDossierPL.x_axis.label_rotation = -45  # ruota di 45 gradi # lblAlgn 

    min_data_row = 2
    max_data_row = ws_dossierpl_filtered.max_row

    data = Reference(ws_dossierpl_filtered, min_col=2, max_col=3, min_row=1, max_row=max_data_row)  # Include header per i titoli
    cats = Reference(ws_dossierpl_filtered, min_col=1, min_row=2, max_row=max_data_row)

    chartDossierPL.add_data(data, titles_from_data=True)
    chartDossierPL.set_categories(cats)
    
    #=======
    # Grafico KPI Cliente DOSSIER PL
    #=======
    # 1. Crea foglio dedicato per grafico eredi
    ws_rintracci_filtered = wb.create_sheet("KPI RINTRACCI GRAFICO")
    ws_rintracci_filtered.sheet_state = 'hidden'
    ws_rintracci_filtered.append(["Cliente", "SLA IN", "SLA OUT"])

    for row in range(2, ws_t.max_row + 1):
        cliente = ws_t.cell(row=row, column=1).value
        sla_in  = ws_t.cell(row=row, column=14).value or 0
        sla_out = ws_t.cell(row=row, column=15).value or 0
        if sla_in > 0 or sla_out > 0:
            ws_rintracci_filtered.append([cliente, sla_in, sla_out])
    
    # 2. Crea grafico usando solo i dati filtrati
    chartRintracci = BarChart()
    chartRintracci.title = "KPI RINTRACCI"
    chartRintracci.y_axis.title = "Numero pratiche"
    chartRintracci.legend.position = "tr"
    chartRintracci.height = 12  # es. 15
    chartRintracci.x_axis.label_rotation = -45  # ruota di 45 gradi # lblAlgn 

    min_data_row = 2
    max_data_row = ws_rintracci_filtered.max_row

    data = Reference(ws_rintracci_filtered, min_col=2, max_col=3, min_row=1, max_row=max_data_row)  # Include header per i titoli
    cats = Reference(ws_rintracci_filtered, min_col=1, min_row=2, max_row=max_data_row)

    chartRintracci.add_data(data, titles_from_data=True)
    chartRintracci.set_categories(cats)
    
    
    
    
    
    
    # ---------------- KPI Grafici
    ws_kpi_graph = wb.create_sheet("KPI Grafici")
    # Posizioniamo il grafico nel foglio, es. cella "B15"
    ws_kpi_graph.add_chart(chart, "B2")
    ws_kpi_graph.add_chart(chartBacc, "B25")
    ws_kpi_graph.add_chart(chartDossierPL, "L2")
    ws_kpi_graph.add_chart(chartRintracci, "L25")
    
    
    
    # 3) Scrivi in memoria
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    #return out
    
    # 4. Scrivi il BytesIO in un file temporaneo .xlsx
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmpfile:
        tmpfile.write(out.read())
        tmpfile_path = tmpfile.name  # es. /tmp/tmpa1b2c3.xlsx    
    
    # 5. Chiama la funzione di patch, che ti restituisce un nuovo file .xlsx
    
    final_xlsx_path = abilita_assi_chart(tmpfile_path)

    # 4. Leggi il file patchato
    with open(final_xlsx_path, 'rb') as f:
        patched_data = f.read()

    # 5. Pulisci i file temporanei
    os.remove(tmpfile_path)
    os.remove(final_xlsx_path)

    # 6. Ritorna il file modificato come download
    response = HttpResponse(
        patched_data,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="quality_report_patch.xlsx"'
    return response