import csv
import datetime, math
from datetime import timedelta
from io import BytesIO
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, LineChart, Reference, Series
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.legend import Legend
from django.core.exceptions import ObjectDoesNotExist
from controlloqualita.services.rules import calcola_punteggio_dynamic

from openpyxl import load_workbook
from openpyxl.styles import Font
import zipfile
from lxml import etree
import tempfile
import os
from django.http import HttpResponse
from controlloqualita.utils import abilita_assi_chart
from collections import defaultdict, OrderedDict



# --- costanti colonne (se proprio le vuoi a indice, altrimenti passa a DictReader) ---
COL_CF_RICH = 13
COL_CF_ATT  = 14
COL_SERV    = 4
COL_CLI     = 2
COL_OP      = 11
COL_DEC     = 20
COL_TEL     = 22
COL_NOTEIN  = 130
COL_INS     = 6
COL_EVA     = 7
COL_SCA     = 134
COL_CALL1   = 85
COL_LAV     = 41
COL_PEN     = 54
COL_BAN     = 70

SERVICE_TO_CATEGORY = {
    'DOSSIER RECUPERO COMPLETO PF': 'DOSSIER_BACC',
    'PERFORMANCE REPORTING PF':      'DOSSIER_BACC',
    'RAPPORTO COMPLETO PLUS':        'DOSSIER_BACC',
    'DOSSIER RECUPERO PF BASE':      'DOSSIER_PL',
    'DOSSIER RECUPERO PF':           'DOSSIER_PL',
    'DOSSIER RECUPERO UGC':          'DOSSIER_PL',
    'RINTRACCIO POSTO LAVORO':       'DOSSIER_PL',
    'RINTRACCIO EREDI':              'EREDI',
    'ACCETTAZIONE EREDI':            'EREDI',
    'RINTRACCIO INDIRIZZO':          'RINTRACCI',
    'CERTIFICAT':                    'CERTIFICATI',
}

def get_categoria_from_servizio(servizio):
    for pattern, cat in SERVICE_TO_CATEGORY.items():
        if pattern in servizio:
            return cat
    return 'N/A'

def get_categoria(servizio):
    for pat, cat in SERVICE_TO_CATEGORY.items():
        if pat in servizio:
            return cat
    return 'N/A'


def parse_aaaammgg(s):
    s = s.strip()
    if len(s) != 8: return None
    try:
        return datetime.date(int(s[:4]), int(s[4:6]), int(s[6:8]))
    except:
        return None


def parse_ddmmyyyy_hhmmss(s):
    s = s.strip()
    if not s: return None
    try:
        day, month, year = s.split()[0].split('/')
        return datetime.date(int(year), int(month), int(day))
    except:
        return None


def calcola_punteggio(SOG):
    cat = SOG['categoria']
    D = SOG['dettagli']
    # decidi quale “condition” applicare
    if cat == 'DOSSIER_BACC':
        if D['decesso_flag'] == 'SI':
            cond = 'decesso'
        elif (D['posto_lavoro'] or D['pensione']) and D['banca']:
            cond = 'lavoro_banca'
        elif (D['posto_lavoro'] or D['pensione']):
            cond = 'lavoro_no_banca'
        elif not (D['posto_lavoro'] or D['pensione'] or D['banca']) \
             and D['telefono'] and 'contattato' in D['note_interne'].lower():
            cond = 'contattato'
        else:
            cond = 'default'

    elif cat == 'DOSSIER_PL':
        # simile a DOSSIER_BACC ma salta P1
        if D['decesso_flag'] == 'SI':
            cond = 'decesso'
        elif (D['posto_lavoro'] or D['pensione']):
            cond = 'lavoro_banca'   # per PL usi sempre P2
        elif D['telefono'] and 'contattato' in D['note_interne'].lower():
            cond = 'contattato'
        else:
            cond = 'default'

    elif cat == 'EREDI':
        cond = 'eredi_chiamato' if D['erede'] else 'eredi_default'

    elif cat == 'RINTRACCI':
        cond = 'rintracci_pos' if (D['residenza_indirizzo'] or D['decesso_flag']=='SI')\
               else 'rintracci_neg'

    else:
        cond = 'default'

    # ora leggi dal DB
    from .models import ScoringRule
    try:
        rule = ScoringRule.objects.get(category=cat, condition=cond)
        return rule.score_letter, rule.score_value
    except ObjectDoesNotExist:
        # fallback se manca la riga
        return ('N2', 3)
    
def calcola_punteggio_old(detail):
    cat = detail['categoria']
    df = detail['dettagli']['decesso_flag']
    lav = detail['dettagli']['posto_lavoro']
    pen = detail['dettagli']['pensione']
    ban = detail['dettagli']['banca']
    tel = detail['dettagli']['telefono']
    note = detail['dettagli']['note_interne'].lower()
    erd = detail['dettagli']['erede']
    res = detail['dettagli']['residenza_indirizzo']

    if cat == 'DOSSIER_BACC':
        if df == 'SI': return 'P2', 10
        if (lav or pen) and ban: return 'P1', 6
        if (lav or pen): return 'P2', 10
        if not (lav or pen or ban) and tel and 'contattato' in note:
            return 'N1', 5
        return 'N2', 3

    if cat == 'DOSSIER_PL':
        if df == 'SI': return 'P2', 10
        if (lav or pen): return 'P2', 10
        if not (lav or pen) and tel and 'contattato' in note:
            return 'N1', 5
        return 'N2', 3

    if cat == 'EREDI':
        return ('P', 10) if erd else ('N', 3)

    if cat == 'RINTRACCI':
        return ('P', 10) if (res or df=='SI') else ('N', 3)

    return ('N2', 3)


def generate_quality_report(csv_path):
    """
    Legge il CSV, aggrega per (CF, cliente, servizio), calcola punteggi e KPI,
    e restituisce un BytesIO con il file Excel pronto.
    """
    # 1) lettura e raggruppamento
    soggetti = {}
    with open(csv_path, newline='', encoding='utf-8') as f:
        reader = csv.reader(f, delimiter=';', quotechar='"')
        next(reader)  # salta header
        for row in reader:
            if not row: continue
            cf = row[COL_CF_RICH].strip()
            cli = row[COL_CLI].strip()
            srv = row[COL_SERV].strip()
            key = (cf, cli, srv)
            if not key in soggetti:
                soggetti[key] = {
                    'rows': [], 'categoria': get_categoria(srv),
                    'cf': cf, 'cliente': cli, 'servizio': srv
                }
            soggetti[key]['rows'].append(row)

    # 2) costruzione dei dettagli e calcolo punteggio + raccolta KPI
    riepilogo = []
    kpi_qual = {}
    kpi_temp = {}

    for (cf, cli, srv), data in soggetti.items():
        rows = data['rows']
        cat  = data['categoria']
        # prendo la prima riga e gli eventuali dati “variabili”
        first = rows[0]
        detail = {
            'categoria': cat,
            'decesso_flag': first[COL_DEC].strip(),
            'telefono': first[COL_TEL].strip(),
            'note_interne': first[COL_NOTEIN].strip(),
            'erede': first[COL_CALL1].strip(),
            'posto_lavoro': bool(first[COL_LAV].strip()),
            'pensione': bool(first[COL_PEN].strip()),
            'banca': bool(first[COL_BAN].strip()),
            'residenza_indirizzo': first[COL_INS]  # o COL_RES
        }
        # puoi fare merge di più righe se ti serve

        # data evasion etc.
        detail['data_evasione'] = parse_aaaammgg(first[COL_EVA])
        detail['data_scadenza'] = parse_ddmmyyyy_hhmmss(first[COL_SCA])
        detail['redattore'] = first[COL_OP].upper()

        # calcolo
        let, num = calcola_punteggio(detail)

        # aggiungo al riepilogo
        riepilogo.append((first[COL_NOM], cf, detail['redattore'],
                          srv, cli, let))

        # KPI qualità
        kpi_qual.setdefault((cli, cat), []).append(num)
        # KPI tempi
        in_time = (detail['data_evasione'] <= detail['data_scadenza']) \
                  if detail['data_evasione'] and detail['data_scadenza'] else None
        kpi_temp.setdefault((cli, cat), []).append(in_time)

    # 3) creo workbook
    wb = Workbook()
    # … (identico al tuo script: foglio Riepilogo, KPI Qualità, KPI tempi, grafici) …
    # QUI build fogli esattamente come nel tuo main(), usando openpyxl.

    # 4) scrivo in BytesIO e restituisco
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# categorie “ufficiali” + la jolly
CATS = ['EREDI', 'DOSSIER_BACC', 'DOSSIER_PL', 'RINTRACCI', 'ALTRO']
SCORE_MAP = {'P1': 6, 'P2': 10, 'N1': 5, 'N2': 3, 'P': 10, 'N': 3}


def generate_quality_report_from_qs(qs):
    """
    Crea un report XLSX:
      • Riepilogo
      • KPI Qualità
      • KPI Tempi
      • KPI Grafici Tempi (bar‑chart)
      • (se un solo cliente) KPI Torta + Grafici Qualità (line‑chart)
    Restituisce un oggetto HttpResponse pronto al download.
    """
    clienti_distinti = list(qs.values_list('ragione_sociale_cliente', flat=True).distinct())
    # ---------- raccolta dati ------------------------------------------------
    dati_riep     = []
    kpi_q_media   = defaultdict(list)     # (cliente,cat) -> [voti]
    kpi_t         = defaultdict(list)     # (cliente,cat) -> [bool]
    qualità_daily = defaultdict(lambda: defaultdict(list))
    #                  ^giorno               ^cat           ^numerico

    for rec in qs:
        servizio = rec.servizio or rec.raw_data.get('Servizio', '')
        cliente  = rec.ragione_sociale_cliente or rec.raw_data.get('RagioneSocialeCliente', '')
        categoria = get_categoria_from_servizio(servizio) or 'ALTRO'

        # ---- punteggio dinamico (lettera/numero)
        p_lettera, p_num = calcola_punteggio_dynamic({
            'categoria': categoria,
            'dettagli': {
                'decesso_flag'        : rec.raw_data.get('DECESSO_FLAG', ''),
                'posto_lavoro'        : bool(rec.raw_data.get('LAVORO_DATORE_RAGIONE_SOCIALE', '').strip()),
                'pensione'            : bool(rec.raw_data.get('PENSIONE_ENTE_RAGIONE_SOCIALE', '').strip()),
                'banca'               : bool(rec.raw_data.get('BANCA_RAGIONE_SOCIALE', '').strip()),
                'telefono'            : rec.raw_data.get('TELEFONO', ''),
                'note_interne'        : rec.raw_data.get('NOTE_INTERNE', ''),
                'erede'               : rec.raw_data.get("CHIAMATO ALL'EREDITA' 1", ''),
                'residenza_indirizzo' : rec.raw_data.get('RESIDENZA_NEW_INDIRIZZO', ''),
            }
        })

        dati_riep.append([
            rec.raw_data.get('NOMINATIVO', ''),
            rec.codice_fiscale_richiesta,
            rec.operatore,
            servizio,
            cliente,
            p_lettera
        ])

        kpi_q_media[(cliente, categoria)].append(p_num)

        di = parse_aaaammgg(rec.raw_data.get('DataInserimento', ''))
        de = parse_aaaammgg(rec.raw_data.get('DataEvasione', ''))
        ds = parse_ddmmyyyy_hhmmss(rec.raw_data.get('DataScadenza', ''))

        kpi_t[(cliente, categoria)].append(None if not (de and ds) else de <= ds)

        # --- per grafico qualità nel tempo
        if di:
            qualità_daily[di][categoria].append(p_num)

    # ---------- workbook -----------------------------------------------------
    wb  = Workbook()
    ws1 = wb.active
    ws1.title = "Riepilogo"
    ws1.append(["Nominativo", "CF", "Redattore", "Servizio", "Cliente", "Punteggio"])
    for r in dati_riep:
        ws1.append(r)

    # ---- KPI Qualità (media 0‑10) ------------------------------------------
    ws_q = wb.create_sheet("KPI Qualità")
    ws_q.append(["Cliente"] + CATS)
    media_cli = defaultdict(dict)
    for (cli, cat), vals in kpi_q_media.items():
        media_cli[cli][cat] = sum(vals) / len(vals)

    order = sorted(media_cli, key=lambda c: sum(media_cli[c].get(cat, 0) for cat in CATS))
    for cli in order:
        ws_q.append([cli] + [round(media_cli[cli].get(cat, 0), 2) for cat in CATS])

    # ---- KPI Tempi (in/out SLA) --------------------------------------------
    ws_t = wb.create_sheet("KPI Tempi")
    suff = ["_numIn", "_numOut", "_%In", "_%Out"]
    ws_t.append(["Cliente"] + [f"{c}{s}" for c in CATS for s in suff])

    for cli in order:
        row = [cli]
        for cat in CATS:
            v = [x for x in kpi_t[(cli, cat)] if x is not None]
            tot = len(v)
            num_in  = sum(v)
            num_out = tot - num_in
            perc_in, perc_out = (round(num_in / tot * 100, 2), round(num_out / tot * 100, 2)) if tot else (0, 0)
            row += [num_in, num_out, perc_in, perc_out]
        ws_t.append(row)

    # ------------------------------------------------------------------
    #  helper: crea e restituisce un bar‑chart con assi visibili
    # ------------------------------------------------------------------
    def _bar_chart_from(wb, ws_src, col_off, title):
        ws_tmp = wb.create_sheet(f"tmp_{title}")   # ← niente state=
        ws_tmp.sheet_state = 'hidden'              # ← lo nascondo dopo

        ws_tmp.append(["Cliente", "SLA IN", "SLA OUT"])
        for r in range(2, ws_src.max_row + 1):
            cli = ws_src.cell(r, 1).value
            sla_in  = ws_src.cell(r, col_off).value or 0
            sla_out = ws_src.cell(r, col_off + 1).value or 0
            if sla_in or sla_out:
                ws_tmp.append([cli, sla_in, sla_out])

        ch = BarChart()
        ch.title = title
        ch.y_axis.title = "Pratiche"
        ch.legend.position = "tr"
        ch.height = 12
        ch.x_axis.label_rotation = -45
        ch.x_axis.visible = True
        ch.y_axis.visible = True

        data = Reference(ws_tmp, min_col=2, max_col=3, min_row=1, max_row=ws_tmp.max_row)
        cats = Reference(ws_tmp, min_col=1, min_row=2, max_row=ws_tmp.max_row)
        ch.add_data(data, titles_from_data=True)
        ch.set_categories(cats)
        return ch        
        
        
        
        

    ws_bar = wb.create_sheet("KPI Grafici Tempi")
    ws_bar.add_chart(_bar_chart_from(wb, ws_t, 2,  "KPI EREDI"),      "B2")
    ws_bar.add_chart(_bar_chart_from(wb, ws_t, 6,  "KPI BACC"),       "B25")
    ws_bar.add_chart(_bar_chart_from(wb, ws_t, 10, "KPI DOSSIER PL"), "L2")
    ws_bar.add_chart(_bar_chart_from(wb, ws_t, 14, "KPI RINTRACCI"),  "L25")

    # ------------------ GRAFICI QUALITÀ (LINE CHART) ------------------
    # Se l’utente ha filtrato un solo cliente, disegna l’andamento
    if len(clienti_distinti) == 1:

        foglio_linee = wb.create_sheet("Grafici Qualità")

        # 1) raccogli tutti i record con data scadenza valida
        dati_cat = {cat: defaultdict(list) for cat in CATS}
        date_set = set()

        for rec in qs:
            cat = get_categoria_from_servizio(rec.servizio)
            if cat not in CATS:          # salta categorie “ALTRO”
                continue
            ds = parse_ddmmyyyy_hhmmss(rec.raw_data.get('DataScadenza', ''))
            if not ds:
                continue
            date_set.add(ds)

            score_letter, score_num = calcola_punteggio_dynamic({
                'categoria': cat,
                'dettagli': {
                    'decesso_flag'      : rec.raw_data.get('DECESSO_FLAG', ''),
                    'posto_lavoro'      : bool(rec.raw_data.get('LAVORO_DATORE_RAGIONE_SOCIALE','').strip()),
                    'pensione'          : bool(rec.raw_data.get('PENSIONE_ENTE_RAGIONE_SOCIALE','').strip()),
                    'banca'             : bool(rec.raw_data.get('BANCA_RAGIONE_SOCIALE','').strip()),
                    'telefono'          : rec.raw_data.get('TELEFONO', ''),
                    'note_interne'      : rec.raw_data.get('NOTE_INTERNE', ''),
                    'erede'             : rec.raw_data.get("CHIAMATO ALL'EREDITA' 1",''),
                    'residenza_indirizzo': rec.raw_data.get('RESIDENZA_NEW_INDIRIZZO',''),
                }})
            dati_cat[cat][ds].append(score_num)

        if date_set:
            # 2) intervalli: dividiamo in 4 parti uguali
            min_d, max_d = min(date_set), max(date_set)
            delta_tot = (max_d - min_d).days
            step = max(1, delta_tot // 4)
            punti = [min_d + timedelta(days=step * i) for i in range(4)]
            punti[-1] = max_d        # garantisci ultimo = max_d

            # 3) intestazione tabella
            foglio_linee.append(["Data"] + CATS)

            # preparo last_val per “carry‑forward”
            last_val = {cat: None for cat in CATS}

            for pt in punti:
                # primo giorno >= pt con dati
                next_dates = sorted(d for d in date_set if d >= pt)
                first_day = next_dates[0] if next_dates else None

                if first_day:
                    lbl = first_day.strftime("%d/%m/%Y")
                else:
                    # non dovrebbe capitare, ma gestiamo
                    lbl = pt.strftime("%d/%m/%Y")

                row = [lbl]
                for cat in CATS:
                    if first_day and dati_cat[cat].get(first_day):
                        # calcola media e aggiorna last_val
                        vals = dati_cat[cat][first_day]
                        media = round(sum(vals)/len(vals), 2)
                        last_val[cat] = media
                    # se non ci sono dati, ri‑usa il valore precedente (carry‑forward)
                    row.append(last_val[cat])
                foglio_linee.append(row)

            # 4) LineChart
            ch = LineChart()
            ch.title = "Trend Qualità (media primo giorno di ogni quarto)"
            ch.y_axis.title = "Indice Qualità"
            ch.x_axis.title = "Data Scadenza"
            ch.y_axis.scaling.min = 0
            ch.y_axis.scaling.max = 10
            ch.height = 12
            ch.width  = 24
            ch.legend.position = "r"
            ch.x_axis.visible = True
            ch.y_axis.visible = True

            data  = Reference(foglio_linee, min_col=2, max_col=len(CATS)+1,
                              min_row=1, max_row=foglio_linee.max_row)
            catsX = Reference(foglio_linee, min_col=1, max_col=1,
                              min_row=2, max_row=foglio_linee.max_row)

            ch.add_data(data, titles_from_data=True)
            ch.set_categories(catsX)

            # etichette valore sui punti
            #for s in ch.series:
            #    s.dLbls = DataLabelList()
            #    s.dLbls.showVal = True
            foglio_linee.add_chart(ch, "B2")    
        
        
    # ---------- output in memoria + patch assi ------------------------------
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(bio.read())
        tmp_path = tmp.name

    patched_path = abilita_assi_chart(tmp_path)

    with open(patched_path, 'rb') as f:
        content = f.read()

    os.remove(tmp_path)
    os.remove(patched_path)

    resp = HttpResponse(
        content,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp['Content-Disposition'] = 'attachment; filename=\"quality_report.xlsx\"'
    return resp
    